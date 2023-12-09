import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
from tkcalendar import DateEntry
import os
import json
import sys
import csv
import openpyxl
import pyperclip
from datetime import datetime,time,timedelta


class FileSelectorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Scraping Data")
        self.root.geometry("500x300")

        self.style = ttk.Style()
        self.style.configure('TButton', font=('Arial', 10), background='#4CAF50')

        self.file_path = tk.StringVar()

        # Set default values for initial and secondary dates and times
        self.set_default_values()

        self.create_widgets()

    def set_default_values(self):
        
        # Set default values for initial date and time
        #initial_date_default = datetime.datetime.today() - datetime.timedelta(days=3)
        #initial_date_default = datetime.datetime(2023, 10, 1)
        #self.initial_date1 = tk.StringVar(value=initial_date_default.strftime('%Y-%m-%d'))
        
        self.initial_date = tk.StringVar(value=(datetime.now()- timedelta(days=3)).strftime('%Y-%m-%d'))
        self.initial_hour = tk.StringVar(value="00")
        self.initial_minute = tk.StringVar(value="01")

        # Set default values for secondary date and time
        self.secondary_date = tk.StringVar(value=datetime.now().strftime('%Y-%m-%d'))
        self.secondary_hour = tk.StringVar(value=datetime.now().strftime('%H'))
        self.secondary_minute = tk.StringVar(value=datetime.now().strftime('%M'))

    def create_widgets(self):
        # File path entry
        label_file = ttk.Label(self.root, text="Select a file:")
        label_file.grid(row=0, column=0, padx=10, pady=10)
        entry_file = ttk.Entry(self.root, textvariable=self.file_path, width=30)
        entry_file.grid(row=0, column=1, padx=10, pady=10)
        browse_button = ttk.Button(self.root, text="Browse", command=self.browse_file)
        browse_button.grid(row=0, column=2, padx=10, pady=10)

        # Initial date entry
        label_initial_date = ttk.Label(self.root, text="Start Date:")
        label_initial_date.grid(row=1, column=0, padx=10, pady=10)
        initial_date_entry = DateEntry(self.root, textvariable=self.initial_date, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        initial_date_entry.grid(row=1, column=1, padx=10, pady=10)

        # Initial time entry
        label_initial_time = ttk.Label(self.root, text="Start Time:")
        label_initial_time.grid(row=2, column=0, padx=10, pady=10)
        initial_hour_spinbox = ttk.Spinbox(self.root, textvariable=self.initial_hour, from_=0, to=23, width=5, format="%02.0f")
        initial_hour_spinbox.grid(row=2, column=1, padx=10, pady=10)
        ttk.Label(self.root, text=" : ").grid(row=2, column=2)
        initial_minute_spinbox = ttk.Spinbox(self.root, textvariable=self.initial_minute, from_=0, to=59, width=5, format="%02.0f")
        initial_minute_spinbox.grid(row=2, column=3, padx=10, pady=10)

        # Secondary date entry
        label_secondary_date = ttk.Label(self.root, text="End Date:")
        label_secondary_date.grid(row=3, column=0, padx=10, pady=10)
        secondary_date_entry = DateEntry(self.root, textvariable=self.secondary_date, width=12, background='darkblue', foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
        secondary_date_entry.grid(row=3, column=1, padx=10, pady=10)

        # Secondary time entry
        label_secondary_time = ttk.Label(self.root, text="End Time:")
        label_secondary_time.grid(row=4, column=0, padx=10, pady=10)
        secondary_hour_spinbox = ttk.Spinbox(self.root, textvariable=self.secondary_hour, from_=0, to=23, width=5, format="%02.0f")
        secondary_hour_spinbox.grid(row=4, column=1, padx=10, pady=10)
        ttk.Label(self.root, text=" : ").grid(row=4, column=2)
        secondary_minute_spinbox = ttk.Spinbox(self.root, textvariable=self.secondary_minute, from_=0, to=59, width=5, format="%02.0f")
        secondary_minute_spinbox.grid(row=4, column=3, padx=10, pady=10)

        # OK Button
        ok_button = ttk.Button(self.root, text="OK", command=self.ok_button_click)
        ok_button.grid(row=5, column=1, padx=10, pady=10)

        self.root.configure(bg='#f0f0f0')

    def browse_file(self):
        file_types = [("Excel files", "*.xlsx;*.xls;*.csv"), ("All files", "*.*")]
        file_path = filedialog.askopenfilename(title="Select a file", filetypes=file_types)

        if file_path:
            _, file_extension = os.path.splitext(file_path)
            valid_extensions = ['.xlsx', '.xls', '.csv']
            if file_extension.lower() in valid_extensions:
                self.file_path.set(file_path)
            else:
                messagebox.showwarning("Invalid File", "Please select a valid Excel or CSV file.")
        else:
            messagebox.showinfo("Info", "No file selected.")

    def ok_button_click(self):
        file_path = self.file_path.get()
        initial_date = self.initial_date.get()
        initial_hour = self.initial_hour.get()
        initial_minute = self.initial_minute.get()
        secondary_date = self.secondary_date.get()
        secondary_hour = self.secondary_hour.get()
        secondary_minute = self.secondary_minute.get()
        data={
        "file_path": file_path,    
        "initial_date"   :  initial_date,
        "initial_hour"   : initial_hour,
        "initial_minute" : initial_minute,
        "secondary_date" : secondary_date,
        "secondary_hour" : secondary_hour,
        "secondary_minute":secondary_minute}
        
        #first_date = datetime.datetime.combine(datetime.datetime.strptime(initial_date, "%Y-%m-%d"), datetime.time(int(initial_hour),int(initial_minute)))
        #end_date = datetime.datetime.combine(datetime.datetime.strptime(secondary_date, "%Y-%m-%d"), datetime.time(int(secondary_hour),int(secondary_minute)))
        #print("Input Values", f"File Path: {file_path}\nInitial Date: {first_date}\nSecondary Date: {end_date}")
        #print(initial_date,initial_hour,initial_minute,secondary_date,secondary_hour,secondary_minute)
    
        serial=json.dumps(data)
        main(serial)


BASE_DIR1 = os.getcwd()
a = BASE_DIR1.split("\\")
BASE_DIR_SPLITED = a[0]+"\\"+a[1]+"\\"+a[2]


# path_csv_raw = BASE_DIR+"\\Desktop\\Book1.csv"
# path_csv_result = BASE_DIR+"\\Desktop\\Book2.csv"
# first_date = "12:14,11.27.2023"
# end_date = "12:14,11.27.2023"

class TempInfo():
    
    #def __init__(self,path_csv_raw,path_csv_result,first_date,end_date,first_time,end_time,type=".csv"):        
    def __init__(self,path_csv_raw,path_csv_result,first_date,end_date,type_input=".xlsx",type_output=".csv"):        
        
        self.type_input = type_input
        
        self.content_scraping = [] # result
        #self.data_time_xlsx_combined = [] # combined the columns 1 & 2 (date and time )
        if type_input == ".csv":
            self.path_csv_raw = BASE_DIR_SPLITED + path_csv_raw + type_input         #"\\Desktop\\Book1.csv"
        elif type_input == ".xlsx":
            #self.path_xlsx_raw = BASE_DIR_SPLITED + path_csv_raw + type_input #when called feom main_cmd func
            self.path_xlsx_raw =  path_csv_raw # when called feom main func (gui)
        self.path_csv_result = BASE_DIR_SPLITED + path_csv_result + type_output   #"\\Desktop\\Book2.csv"
        
        self.first_date = first_date  #"12:14,11.27.2023"
        self.end_date = end_date      #"12:14,11.27.2023"
       
        if self.type_input == ".csv":
            self.content_before = self.read_from_csv()
            self.content_after = self.proccess_my_csv_data()
        elif self.type_input == ".xlsx":   
            self.content_before = self.read_from_xlsx()
            self.content_after = None

    def update_content(self):
        if self.type_input == ".csv":
            self.content_before = self.read_from_csv()
            self.content_after = self.proccess_my_csv_data()
        elif self.type_input == ".xlsx":
            self.content_before = self.read_from_xlsx()
            self.content_after = None
        
    def read_from_csv(self):
        with open(self.path_csv_raw,"r") as f:
           content = f.read()
        return content
           
    def read_from_xlsx(self):
        workbook = openpyxl.load_workbook(self.path_xlsx_raw)
        sheet = workbook.active
        temp_list_of_cell = []
        for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                temp_list_of_cell.append(cell)  
        return temp_list_of_cell
    
    
    def proccess_my_csv_data(self):
        content = self.content_before
        data_proccessed = content[3:].replace("\n",",").replace("\"","").split(",")[:-1] #remove \n then " from data and seperate the string to list with , and ignore the last item of list becuase it's ''         
        return data_proccessed

    # def proccess_my_xlsx_data(self):
    #     self.data_time_xlsx_combined = []
    #     content = self.content_before # the content before is processed by the openpyxl and no need to process it again
    #     for i in range(0,len(content),3):
    #         if content[i].value != None and content[i+1].value != None: 
    #             self.data_time_xlsx_combined.append(datetime.combine(content[i].value,content[i+1].value))

    
    def scrape_my_data(self):
        #print(self.content_after)
        self.content_scraping  = []
        #print(len(self.content_after))
        if self.type_input == ".csv":
            for i in range(0,len(self.content_after),3):
                #print(i)
                #_time = self.content_after[i].split(":") 
                _time = self.content_after[i+1] 
                _date = self.content_after[i]
                #print(_date.strip()+" "+_time)
                #_tstruct = time(int(_time[0]),int(_time[1]))
                try:
                    _dstruct = datetime.strptime(_date.strip()+" "+_time.strip() ,"%m/%d/%Y %H:%M")
                except ValueError:
                    continue
                #print(_dstruct)
                #if self.first_date <= _dstruct and self.first_time <= _tstruct and self.end_date >= _dstruct and self.end_time >= _tstruct :
                if self.first_date <= _dstruct and self.end_date >= _dstruct:
                        self.content_scraping.append(self.content_after[i])
                        self.content_scraping.append(self.content_after[i+1])
                        self.content_scraping.append(self.content_after[i+2])
                                        
                elif self.end_date < _dstruct:
                    break
        if self.type_input == ".xlsx":
            content = self.content_before # the content before is processed by the openpyxl and no need to process it again
            #print(content)
            for i in range(3,len(content),3):
                _dstruct1 = datetime(year=2000,month=1,day=1)
                if content[i].value != None and content[i+1].value != None:
                    #print(content[i].value , content[i+1].value)
                    _dstruct1 = datetime.combine(content[i].value,content[i+1].value)
                    
                    
                    if self.first_date <= _dstruct1 and self.end_date >= _dstruct1:
                            #print(i," : ",_dstruct1)
                            index = i
                            self.content_scraping.append(self.content_before[index].value)
                            self.content_scraping.append(self.content_before[index+1].value)
                            self.content_scraping.append(self.content_before[index+2].value)
                            
            
                elif self.end_date < _dstruct1:
                    break
            #print(self.content_scraping)
    
    def write_to_csv(self):
        if os.path.exists(self.path_csv_result): 
            os.remove(self.path_csv_result) 
        data_proccessed = self.content_scraping
        with open(self.path_csv_result,"a",encoding='UTF8',newline="") as f:
            a = csv.writer(f)
            for i in range(0,len(data_proccessed),3):  
                a.writerow([str(data_proccessed[i]).split(" ")[0].replace("-","/"),":".join(str(data_proccessed[i+1]).split(":")[0:2]),str(data_proccessed[i+2])])
    
    def show_my_data(self,before=False,after=False):
        if before == True:
            print("### content before processed ###")
            print(self.content_before)
        
        if after == True:
            if self.type_input != ".xlsx":    
                print("### content after processed ###")
                print(self.content_after)
            

    def delete_my_data(self):
        del self.content_after
        del self.content_before
        
#print(content)
def data_input(path = False):
    def get_path():
        path_xlsx_raw = input("path_your_file_xlsx (defult=\\Desktop\\data.xlsx):")
        if path_xlsx_raw == "":
            path_xlsx_raw = "\\Desktop\\data"
        else:
            path_xlsx_raw = path_xlsx_raw.split(".")[0]
            
        path_xlsx_result = input("path_csv_result (defult=\\Desktop\\output.csv):")
        if path_xlsx_result == "":
            path_xlsx_result = "\\Desktop\\output"
        else:
            path_xlsx_result = path_xlsx_raw.split(".")[0]
        return path_xlsx_raw,path_xlsx_result
    
    def date_and_time():
        
        first_date = input("first_date as format:month/day/year (example=11/24/2023, defualt = three days ago) :")
        first_time = input("first_time as format:hour/minute    (example=12:45 , defualt=00:01) :")
        end_date =   input("\nend_date as format:month/day/year (example=12/27/2023, defualt = today) :")
        end_time =   input("end_time as format:hour/minute      (example=13:15 , defualt = now) :")
        
        now_datetime = datetime.today()
        
        if first_time == "":
            first_time = "00:01"
        if end_time == "":
            end_time = str(now_datetime.hour)+":"+str(now_datetime.minute)
        if first_date == "":
            tow_day_ago = now_datetime - timedelta(days=3)
            first_date = str(tow_day_ago.month)+"/"+str(tow_day_ago.day)+"/"+str(tow_day_ago.year)
        if end_date == "":
            end_date = str(now_datetime.month)+"/"+str(now_datetime.day)+"/"+str(now_datetime.year)
 
        return first_date.strip(),first_time.strip(),end_date.strip(),end_time.strip()
   
    # type = print("type of file (defult = .csv):")
    # if type == "":
    #     type = ".csv"
   
    path_csv_raw,path_csv_result = None,None
    if path == True:
        path_csv_raw,path_csv_result = get_path()
    
    
    #while True:
        
    first_date,first_time,end_date,end_time = date_and_time()
    
    first_date = datetime.strptime(first_date+" "+first_time , "%m/%d/%Y %H:%M")
    end_date = datetime.strptime(end_date+" "+end_time, "%m/%d/%Y %H:%M")
    #first_time = time(int(first_time.split(":")[0]),int(first_time.split(":")[1]))
    #end_time = time(int(end_time.split(":")[0]) ,int(end_time.split(":")[1]))
        #except:
        #    print("\n\n!! ERROR !!\n please try to correct data and time input\n")        
        #else:
        #    break
    #return path_csv_raw,path_csv_result,first_date,first_time,end_date,end_time
    return path_csv_raw,path_csv_result,first_date,end_date


def main_cmd():    
    #path_csv_raw,path_csv_result,first_date,first_time,end_date,end_time = data_input(path=True)
    path_csv_raw,path_csv_result,first_date,end_date = data_input(path=True)
    print(path_csv_raw,path_csv_result)
    #main_object = TempInfo(path_csv_raw,path_csv_result,first_date,end_date,first_time,end_time)     
    main_object = TempInfo(path_csv_raw,path_csv_result,first_date,end_date)     
    
    dict_of_action = {#1:"print data before processing",
                          #2:"print data after processing",
                          3:"inter the new time to scraping",
                          4:"update the content and file",
                          5:"print input time and date",
                          6:"scrap and return data",
                          7:"write data to result file",
                          10:"nothing and close the progrom",}
    flag = 0
    while True:
        
        print()
        if flag == 0:
            for i in list(dict_of_action.items()):
                print("\t",i[0]," : ",i[1],sep="")
                flag = 1
        #input()
        #try:
        try:
            action = int(input("please inter the command to do:"))
        except:
            print("!! You entered an invalid entry. Please enter only the numbers 3, 4, 5, 6, 7, and 10 ")
            continue
        print()
        flag = 0
        # if action == 1:
        #     main_object.show_my_data(before=True)
        # elif action == 2:
        #     main_object.show_my_data(after=True)
        if action == 3:
            #_,_,first_date,first_time,end_date,end_time=data_input(path=False)
            _,_,first_date,end_date=data_input(path=False)
            main_object.first_date = first_date
            main_object.end_date = end_date
            
        elif action == 4:
            main_object.update_content()
            print("update Done!")
            
            #main_object.show_my_data(before=True,after=True)
        elif action == 5:
            print("first_date:{}\nend_date:{}\n".format(first_date,end_date))
        elif action == 6:
            main_object.scrape_my_data()
            for i in range(0,len(main_object.content_scraping),3):
                print(main_object.content_scraping[i]," ",main_object.content_scraping[i+1]," ",main_object.content_scraping[i+2])
        elif action == 7:
            main_object.write_to_csv()   
            print("writed data to output.csv file in Desktop Done!")
        #except:
         #   continue
        else:
            if action == 10:
                quit()
                sys.exit()
     

def show_error_popup(error_message):
    error_popup = tk.Tk()
    error_popup.title("Error")
    
    error_label = tk.Label(error_popup, text=f"\u2716 An error occurred:\n{error_message}", font=("Helvetica", 22), fg="red")
    error_label.pack(padx=10, pady=10)

    ok_button = tk.Button(error_popup, text="OK", command=error_popup.destroy)
    ok_button.pack(pady=10)

    error_popup.mainloop()     

def copy_to_clipboard(content):
    pyperclip.copy(content)

def show_text_window(result_to_copy):
    root = tk.Tk()
    root.title("نتیجه")

    text_widget = tk.Text(root, height=10, width=50)
    text_widget.insert(tk.END, result_to_copy)
    text_widget.pack(padx=20, pady=10)
    copy_button = tk.Button(root, text="Copy to Clipboard", command=lambda: copy_to_clipboard(result_to_copy))
    copy_button.pack(pady=10)

    root.mainloop()
     
def main(data):

    # The first command-line argument is the JSON data passed from the main program
    json_data = data

    # Convert the JSON string to a dictionary
    received_data = json.loads(json_data)
    first_date = datetime.combine(datetime.strptime(received_data["initial_date"], "%Y-%m-%d"), time(int(received_data["initial_hour"]),int(received_data["initial_minute"])))
    end_date = datetime.combine(datetime.strptime(received_data["secondary_date"], "%Y-%m-%d"), time(int(received_data["secondary_hour"]),int(received_data["secondary_minute"])))
    Input_File_Path= received_data['file_path']
    #print(Input_File_Path)
    #print("Input Values", f"File Path: {received_data['file_path']}\nInitial Date: {first_date}\nSecondary Date: {end_date}")
    main_object = TempInfo(Input_File_Path,"\\Desktop\\output",first_date,end_date)     
    main_object.scrape_my_data()
    result_to_copy = ""
    for i in range(0,len(main_object.content_scraping),3):
        result_to_copy += str(main_object.content_scraping[i]).split(" ")[0].replace("-","/")+","+":".join(str(main_object.content_scraping[i+1]).split(":")[0:2])+","+str(main_object.content_scraping[i+2])+"\n"
    
    main_object.write_to_csv()
    show_text_window(result_to_copy)
       
    #print("writed data to output.csv file in Desktop Done!")
    # Perform the desired function with the received data
    #print(f"Received data from the main program: {received_data}")



if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = FileSelectorApp(root)
        root.mainloop()
    except Exception as e:
        show_error_popup(str(e))