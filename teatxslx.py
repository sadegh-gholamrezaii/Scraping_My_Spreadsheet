# import openpyxl as xl
# with xl.load_workbook("C:\\Users\\sadegh\\Desktop\\file.xlsx") as f:
#     sheet = f.active 
#     print(sheet.max_row)

import openpyxl
from datetime import datetime
file_path = 'C:\\Users\\sadegh\\Desktop\\file.xlsx'

# باز کردن فایل
workbook = openpyxl.load_workbook(file_path)

# انتخاب شیت
sheet = workbook.active
a = []
# خواندن محتویات
for row in sheet.iter_rows(min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
    for cell in row:
        print(cell.value, end='\t')
        a.append(cell)  
    print()

date1 = 84
time1 = 85
print(a[date1].value,a[time1].value)
if a[date1].value != None and a[time1].value != None: 
    print(datetime.combine(a[date1].value,a[time1].value))
# بستن فایل به صورت دستی
workbook.close()